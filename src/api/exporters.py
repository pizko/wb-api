from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from api.models import CatalogItem, Group, GroupItemLink, NewItem


def export_catalog_xlsx(path: Path, items: list[CatalogItem]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "catalog"
    sheet.append(
        [
            "id",
            "article",
            "name",
            "discount",
            "price_base",
            "price",
            "qty",
            "image_url",
            "description",
            "material",
            "country",
            "producer",
            "vibration",
            "group_name",
            "barcode",
            "brief",
            "outlet",
            "is_anticrisis",
            "rrc",
            "qty_hr",
        ]
    )
    for item in items:
        sheet.append(
            [
                item.id,
                item.article,
                item.name,
                item.discount,
                item.price_base,
                item.price,
                item.qty,
                item.image_url,
                item.description,
                item.material,
                item.country,
                item.producer,
                item.vibration,
                item.group_name,
                item.barcode,
                item.brief,
                item.outlet,
                item.is_anticrisis,
                item.rrc,
                item.qty_hr,
            ]
        )
    _autosize(sheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def export_groups_xlsx(path: Path, groups: list[Group], links: list[GroupItemLink] | None = None) -> Path:
    workbook = Workbook()
    groups_sheet = workbook.active
    groups_sheet.title = "groups"
    groups_sheet.append(["id", "parent_id", "name"])
    for group in groups:
        groups_sheet.append([group.id, group.parent_id, group.name])
    _autosize(groups_sheet)

    if links is not None:
        links_sheet = workbook.create_sheet("group_links")
        links_sheet.append(["item_id", "group_id"])
        for link in links:
            links_sheet.append([link.item_id, link.group_id])
        _autosize(links_sheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def export_new_items_xlsx(path: Path, items: list[NewItem]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "new_items"
    sheet.append(
        [
            "id",
            "article",
            "name",
            "price",
            "qty",
            "image_url",
            "description",
            "discount",
            "material",
            "country",
            "producer",
            "vibration",
            "arrival_date",
        ]
    )
    for item in items:
        sheet.append(
            [
                item.id,
                item.article,
                item.name,
                item.price,
                item.qty,
                item.image_url,
                item.description,
                item.discount,
                item.material,
                item.country,
                item.producer,
                item.vibration,
                item.arrival_date,
            ]
        )
    _autosize(sheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def export_marketplace_feed_xlsx(
    path: Path,
    items: list[CatalogItem],
    groups: list[Group],
    links: list[GroupItemLink],
) -> Path:
    group_map = {group.id: group for group in groups}
    item_to_groups: dict[int, list[str]] = {}
    for link in links:
        group = group_map.get(link.group_id)
        if not group:
            continue
        item_to_groups.setdefault(link.item_id, []).append(group.name)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "marketplace_feed"
    sheet.append(
        [
            "supplier_item_id",
            "offer_id",
            "article",
            "name",
            "brief",
            "description",
            "price",
            "price_base",
            "rrc",
            "qty",
            "qty_hr",
            "discount",
            "barcode",
            "image_url",
            "material",
            "country",
            "producer",
            "group_name",
            "group_tree_candidates",
            "is_outlet",
            "is_anticrisis",
            "vibration",
        ]
    )
    for item in items:
        sheet.append(
            [
                item.id,
                _build_offer_id(item),
                item.article,
                item.name,
                item.brief,
                item.description,
                item.price,
                item.price_base,
                item.rrc,
                item.qty,
                item.qty_hr,
                item.discount,
                item.barcode,
                item.image_url,
                item.material,
                item.country,
                item.producer,
                item.group_name,
                " | ".join(sorted(item_to_groups.get(item.id, []))),
                item.outlet,
                item.is_anticrisis,
                item.vibration,
            ]
        )
    _autosize(sheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def export_ozon_products_draft_xlsx(path: Path, rows: list[dict[str, str]]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ozon_products_draft"
    sheet.append(
        [
            "offer_id",
            "name",
            "description",
            "barcode",
            "price_rrc",
            "image_url",
            "supplier_group_name",
            "supplier_group_tree",
            "ozon_category_id",
            "ozon_type_id",
            "ozon_name",
            "brand",
            "vat",
            "images_json",
            "attributes_json",
            "comment",
        ]
    )
    for row in rows:
        image_url = (row.get("image_url") or "").strip()
        sheet.append(
            [
                row.get("offer_id", ""),
                row.get("name", ""),
                row.get("description", ""),
                row.get("barcode", ""),
                row.get("rrc", ""),
                image_url,
                row.get("group_name", ""),
                row.get("group_tree_candidates", ""),
                "",
                "",
                row.get("name", ""),
                row.get("producer", ""),
                "0",
                json.dumps([image_url], ensure_ascii=False) if image_url else "[]",
                "[]",
                "Заполни ozon_category_id, brand и attributes_json перед импортом",
            ]
        )
    _autosize(sheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def export_wb_products_draft_xlsx(path: Path, rows: list[dict[str, str]]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "wb_products_draft"
    sheet.append(
        [
            "vendorCode",
            "title",
            "description",
            "brand",
            "price_rrc",
            "length",
            "width",
            "height",
            "weightBrutto",
            "barcode",
            "images_json",
            "supplier_group_name",
            "supplier_group_tree",
            "wb_parent_id",
            "wb_subject_id",
            "wb_subject_name",
            "characteristics_json",
            "comment",
        ]
    )
    for row in rows:
        image_url = (row.get("image_url") or "").strip()
        sheet.append(
            [
                row.get("offer_id", ""),
                row.get("name", ""),
                row.get("description", ""),
                row.get("producer", ""),
                row.get("rrc", ""),
                "",
                "",
                "",
                "",
                row.get("barcode", ""),
                json.dumps([image_url], ensure_ascii=False) if image_url else "[]",
                row.get("group_name", ""),
                row.get("group_tree_candidates", ""),
                "",
                "",
                "",
                "[]",
                "Заполни wb_parent_id, wb_subject_id, характеристики и габариты перед импортом",
            ]
        )
    _autosize(sheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def export_wb_mapping_draft_xlsx(path: Path, rows: list[dict[str, str]], wb_subjects_payload: dict) -> Path:
    workbook = Workbook()

    mapping_sheet = workbook.active
    mapping_sheet.title = "supplier_to_wb_mapping"
    mapping_sheet.append(
        [
            "supplier_group_name",
            "supplier_group_tree",
            "items_count",
            "sample_offer_ids",
            "wb_parent_id",
            "wb_subject_id",
            "wb_subject_name",
            "notes",
        ]
    )

    grouped: dict[tuple[str, str], list[dict[str, str]]] = {}
    for row in rows:
        key = ((row.get("group_name") or "").strip(), (row.get("group_tree_candidates") or "").strip())
        grouped.setdefault(key, []).append(row)

    for (group_name, group_tree), group_rows in sorted(grouped.items(), key=lambda item: item[0]):
        sample_offer_ids = ", ".join(
            sorted({row.get("offer_id", "").strip() for row in group_rows if row.get("offer_id", "").strip()})[:5]
        )
        mapping_sheet.append(
            [
                group_name,
                group_tree,
                len(group_rows),
                sample_offer_ids,
                5038,
                "",
                "",
                "",
            ]
        )
    _autosize(mapping_sheet)

    subjects_sheet = workbook.create_sheet("wb_subjects")
    subjects_sheet.append(["parentID", "parentName", "subjectID", "subjectName"])
    for item in wb_subjects_payload.get("data", []):
        subjects_sheet.append(
            [
                item.get("parentID", ""),
                item.get("parentName", ""),
                item.get("subjectID", ""),
                item.get("subjectName", ""),
            ]
        )
    _autosize(subjects_sheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def autofill_wb_mapping_draft_xlsx(path: Path) -> Path:
    workbook = load_workbook(path)
    if "supplier_to_wb_mapping" not in workbook.sheetnames or "wb_subjects" not in workbook.sheetnames:
        raise RuntimeError("В файле нет листов supplier_to_wb_mapping или wb_subjects")

    mapping_sheet = workbook["supplier_to_wb_mapping"]
    subjects_sheet = workbook["wb_subjects"]

    subject_lookup: dict[str, tuple[int, str]] = {}
    for row in subjects_sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[2] is None or row[3] is None:
            continue
        subject_lookup[str(row[3]).strip().lower()] = (int(row[2]), str(row[3]).strip())

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in mapping_sheet[1]]
    index = {name: idx for idx, name in enumerate(headers)}

    filled = 0
    for excel_row in mapping_sheet.iter_rows(min_row=2):
        group_name = str(excel_row[index["supplier_group_name"]].value or "").strip().lower()
        group_tree = str(excel_row[index["supplier_group_tree"]].value or "").strip().lower()
        wb_subject_id_cell = excel_row[index["wb_subject_id"]]
        wb_subject_name_cell = excel_row[index["wb_subject_name"]]
        notes_cell = excel_row[index["notes"]]
        if wb_subject_id_cell.value:
            continue
        matched = _match_wb_subject(group_name, group_tree, subject_lookup)
        if not matched:
            continue
        wb_subject_id_cell.value = matched[0]
        wb_subject_name_cell.value = matched[1]
        notes_cell.value = "autofilled"
        filled += 1

    workbook.save(path)
    print(f"Автозаполнено строк WB mapping: {filled}")
    return path


def apply_wb_mapping_to_products_draft_xlsx(mapping_path: Path, products_path: Path) -> Path:
    mapping_workbook = load_workbook(mapping_path, read_only=True)
    products_workbook = load_workbook(products_path)

    mapping_sheet = mapping_workbook["supplier_to_wb_mapping"]
    products_sheet = products_workbook.active

    mapping_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in mapping_sheet[1]]
    mapping_index = {name: idx for idx, name in enumerate(mapping_headers)}

    products_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in products_sheet[1]]
    products_index = {name: idx for idx, name in enumerate(products_headers)}

    mapping_lookup: dict[tuple[str, str], tuple[int, str, int]] = {}
    for row in mapping_sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        group_name = str(row[mapping_index["supplier_group_name"]] or "").strip()
        group_tree = str(row[mapping_index["supplier_group_tree"]] or "").strip()
        parent_id = row[mapping_index["wb_parent_id"]]
        subject_id = row[mapping_index["wb_subject_id"]]
        subject_name = str(row[mapping_index["wb_subject_name"]] or "").strip()
        if not subject_id:
            continue
        mapping_lookup[(group_name, group_tree)] = (int(parent_id or 5038), int(subject_id), subject_name)

    applied = 0
    for row in products_sheet.iter_rows(min_row=2):
        group_name = str(row[products_index["supplier_group_name"]].value or "").strip()
        group_tree = str(row[products_index["supplier_group_tree"]].value or "").strip()
        matched = mapping_lookup.get((group_name, group_tree))
        if not matched:
            continue
        row[products_index["wb_parent_id"]].value = matched[0]
        row[products_index["wb_subject_id"]].value = matched[1]
        row[products_index["wb_subject_name"]].value = matched[2]
        applied += 1

    products_workbook.save(products_path)
    print(f"Применено строк маппинга к WB draft: {applied}")
    return products_path


def collect_wb_subject_ids_from_products_draft(path: Path) -> list[int]:
    workbook = load_workbook(path, read_only=True)
    sheet = workbook.active
    headers = [str(cell).strip() if cell is not None else "" for cell in next(sheet.iter_rows(values_only=True))]
    index = {name: idx for idx, name in enumerate(headers)}
    subject_ids: set[int] = set()
    for row in sheet.iter_rows(values_only=True):
        value = row[index["wb_subject_id"]] if index.get("wb_subject_id") is not None and len(row) > index["wb_subject_id"] else None
        try:
            if value:
                subject_ids.add(int(value))
        except (TypeError, ValueError):
            continue
    return sorted(subject_ids)


def autofill_wb_products_characteristics_xlsx(products_path: Path, characteristics_dir: Path) -> Path:
    workbook = load_workbook(products_path)
    sheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    index = {name: idx for idx, name in enumerate(headers)}

    updated_rows = 0
    for row in sheet.iter_rows(min_row=2):
        subject_id = _safe_int_cell(row[index["wb_subject_id"]].value if "wb_subject_id" in index else None)
        if not subject_id:
            continue

        char_path = characteristics_dir / f"subject_characteristics_{subject_id}.json"
        if not char_path.exists():
            continue
        characteristics_payload = json.loads(char_path.read_text(encoding="utf-8"))
        available = _extract_wb_characteristics(characteristics_payload)
        if not available:
            continue

        current_json = str(row[index["characteristics_json"]].value or "").strip()
        if current_json and current_json != "[]":
            continue

        draft_row = {name: str(row[idx].value or "").strip() for name, idx in index.items()}
        filled = _build_wb_characteristics_from_supplier(draft_row, available)
        if not filled:
            continue
        row[index["characteristics_json"]].value = json.dumps(filled, ensure_ascii=False)
        comment_cell = row[index["comment"]]
        existing_comment = str(comment_cell.value or "").strip()
        comment_cell.value = (existing_comment + " | " if existing_comment else "") + "characteristics_autofilled"
        updated_rows += 1

    workbook.save(products_path)
    print(f"Автозаполнено characteristics_json в WB draft: {updated_rows}")
    return products_path


def load_marketplace_feed_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    result: list[dict[str, str]] = []
    for row in rows[1:]:
        payload: dict[str, str] = {}
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else None
            payload[header] = "" if value is None else str(value)
        result.append(payload)
    return result


def load_ozon_products_draft_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    result: list[dict[str, str]] = []
    for row in rows[1:]:
        payload: dict[str, str] = {}
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else None
            payload[header] = "" if value is None else str(value)
        result.append(payload)
    return result


def load_wb_products_draft_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    result: list[dict[str, str]] = []
    for row in rows[1:]:
        payload: dict[str, str] = {}
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else None
            payload[header] = "" if value is None else str(value)
        result.append(payload)
    return result


def load_json_file(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _match_wb_subject(
    group_name: str,
    group_tree: str,
    subject_lookup: dict[str, tuple[int, str]],
) -> tuple[int, str] | None:
    text = f"{group_name} | {group_tree}"
    rules = [
        (("вибропул", "яичк"), "вибраторы"),
        (("вибратор", "wand"), "вибраторы"),
        (("мастурбатор",), "мастурбаторы мужские"),
        (("фалло", "реалистич"), "фаллоимитаторы"),
        (("фаллопротез",), "фаллопротезы"),
        (("насадк", "вибратор"), "насадки для вибраторов"),
        (("насадк", "мастурбатор"), "насадки на мастурбатор"),
        (("насадк", "страпон"), "насадки на страпон"),
        (("насадк", "член"), "насадки на член"),
        (("лубрикант",), "лубриканты"),
        (("крем", "мужчин"), "лубриканты"),
        (("стимулирующ", "гель"), "лубриканты"),
        (("анальн", "пробк"), "анальные пробки"),
        (("анальн", "шарик"), "анальные шарики"),
        (("анальн", "бусы"), "анальные бусы"),
        (("вагинальн", "шарик"), "вагинальные шарики"),
        (("помп",), "вакуумные помпы эротик"),
        (("вакуумно-волнов",), "вакуумно-волновые стимуляторы"),
        (("секс кукл",), "секс куклы"),
        (("электростимул",), "электростимуляторы"),
        (("аксессуар", "игруш"), "аксессуары для игрушек эротик"),
    ]
    for keywords, subject_name in rules:
        if all(keyword in text for keyword in keywords):
            match = subject_lookup.get(subject_name)
            if match:
                return match
    return None


def _extract_wb_characteristics(payload: dict) -> list[dict]:
    if isinstance(payload, dict):
        data = payload.get("data")
        if isinstance(data, list):
            return [item for item in data if isinstance(item, dict)]
    return []


def _build_wb_characteristics_from_supplier(draft_row: dict[str, str], available: list[dict]) -> list[dict]:
    result: list[dict] = []
    name_map: dict[str, dict] = {}
    for item in available:
        name = str(item.get("name") or item.get("charcName") or "").strip().lower()
        if name:
            name_map[name] = item

    def add_if_present(possible_names: tuple[str, ...], value: str) -> None:
        clean = value.strip()
        if not clean:
            return
        for candidate in possible_names:
            meta = name_map.get(candidate.lower())
            if meta:
                result.append(_format_wb_characteristic(meta, clean))
                return

    add_if_present(("Бренд",), draft_row.get("brand", ""))
    add_if_present(("Страна изготовитель", "Страна производства"), draft_row.get("country", ""))
    add_if_present(("Материал изделия", "Материал"), draft_row.get("description", "") if "водная основа" in draft_row.get("description", "").lower() else draft_row.get("supplier_group_name", ""))
    add_if_present(("Состав",), draft_row.get("description", "") if "состав" in draft_row.get("description", "").lower() else "")
    add_if_present(("Комплектация",), draft_row.get("title", ""))
    add_if_present(("Название модели", "Модель"), draft_row.get("vendorCode", ""))

    color = _guess_color(draft_row.get("title", ""))
    add_if_present(("Цвет",), color)

    volume = _guess_volume_ml(draft_row.get("title", ""))
    add_if_present(("Объем", "Объем, мл", "Объём"), volume)

    return _dedupe_characteristics(result)


def _format_wb_characteristic(meta: dict, value: str) -> dict:
    charc_id = meta.get("charcID") or meta.get("id")
    name = meta.get("name") or meta.get("charcName") or ""
    return {
        "id": charc_id,
        "name": name,
        "value": value,
    }


def _dedupe_characteristics(items: list[dict]) -> list[dict]:
    seen: set[tuple[str, str]] = set()
    result: list[dict] = []
    for item in items:
        key = (str(item.get("id", "")), str(item.get("value", "")))
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def _safe_int_cell(value) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return 0


def _guess_color(title: str) -> str:
    text = title.lower()
    colors = {
        "беж": "бежевый",
        "черн": "черный",
        "бел": "белый",
        "красн": "красный",
        "розов": "розовый",
        "син": "синий",
        "фиолет": "фиолетовый",
        "зелен": "зеленый",
        "gold": "золотой",
        "сереб": "серебристый",
    }
    for needle, color in colors.items():
        if needle in text:
            return color
    return ""


def _guess_volume_ml(title: str) -> str:
    import re

    match = re.search(r"(\d+(?:[.,]\d+)?)\s*мл", title.lower())
    if not match:
        return ""
    return match.group(1).replace(",", ".")


def save_json(path: Path, payload: dict | list) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def export_ozon_sync_report_xlsx(
    path: Path,
    *,
    prices_results: list[dict],
    stocks_results: list[dict],
    metadata: dict[str, str | int],
) -> Path:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["metric", "value"])
    for key, value in metadata.items():
        summary_sheet.append([key, value])
    summary_sheet.append(["prices_total", len(prices_results)])
    summary_sheet.append(["prices_updated", sum(1 for row in prices_results if row.get("updated"))])
    summary_sheet.append(["prices_with_errors", sum(1 for row in prices_results if row.get("error_codes"))])
    summary_sheet.append(["stocks_total", len(stocks_results)])
    summary_sheet.append(["stocks_updated", sum(1 for row in stocks_results if row.get("updated"))])
    summary_sheet.append(["stocks_with_errors", sum(1 for row in stocks_results if row.get("error_codes"))])
    for label, count in _count_error_codes(prices_results, prefix="prices").items():
        summary_sheet.append([label, count])
    for label, count in _count_error_codes(stocks_results, prefix="stocks").items():
        summary_sheet.append([label, count])
    _autosize(summary_sheet)

    prices_sheet = workbook.create_sheet("prices")
    _append_ozon_result_rows(prices_sheet, prices_results)

    stocks_sheet = workbook.create_sheet("stocks")
    _append_ozon_result_rows(stocks_sheet, stocks_results)

    prices_errors_sheet = workbook.create_sheet("prices_errors")
    _append_ozon_result_rows(
        prices_errors_sheet,
        [row for row in prices_results if row.get("error_codes")],
    )

    stocks_errors_sheet = workbook.create_sheet("stocks_errors")
    _append_ozon_result_rows(
        stocks_errors_sheet,
        [row for row in stocks_results if row.get("error_codes")],
    )

    for error_code in ("NOT_FOUND", "NOT_FOUND_ERROR", "PRODUCT_IS_NOT_CREATED", "NOT_PASS_MODERATION"):
        error_rows = [
            row
            for row in prices_results + stocks_results
            if error_code in str(row.get("error_codes", ""))
        ]
        sheet = workbook.create_sheet(_sheet_name_for_error(error_code))
        _append_ozon_result_rows(sheet, error_rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _autosize(sheet) -> None:
    for column in sheet.columns:
        values = [len(str(cell.value)) for cell in column if cell.value is not None]
        width = min(max(values, default=10) + 2, 80)
        sheet.column_dimensions[column[0].column_letter].width = width


def _build_offer_id(item: CatalogItem) -> str:
    value = item.article.strip() if item.article else str(item.id)
    return value[:100]


def _append_ozon_result_rows(sheet, rows: list[dict]) -> None:
    headers = [
        "batch",
        "warehouse_id",
        "product_id",
        "offer_id",
        "updated",
        "error_codes",
        "error_messages",
        "warning_codes",
        "warning_messages",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    _autosize(sheet)


def _count_error_codes(rows: list[dict], *, prefix: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        raw_codes = str(row.get("error_codes", "")).strip()
        if not raw_codes:
            continue
        for code in [part.strip() for part in raw_codes.split("|") if part.strip()]:
            counts[f"{prefix}_{code}"] = counts.get(f"{prefix}_{code}", 0) + 1
    return counts


def _sheet_name_for_error(error_code: str) -> str:
    mapping = {
        "NOT_FOUND": "not_found",
        "NOT_FOUND_ERROR": "not_found_error",
        "PRODUCT_IS_NOT_CREATED": "not_created",
        "NOT_PASS_MODERATION": "not_moderated",
    }
    return mapping.get(error_code, error_code.lower()[:31])
